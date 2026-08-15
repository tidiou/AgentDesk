from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.schemas.uat import UATTestCase

HEADER_FILL = PatternFill(start_color="2C2F36", end_color="2C2F36", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("ID", "id", 10),
    ("Requirement", "requirement_ref", 15),
    ("Title", "title", 25),
    ("Preconditions", "preconditions", 25),
    ("Steps", "steps", 40),
    ("Expected Result", "expected_result", 30),
    ("Priority", "priority", 10),
]


def build_uat_excel(test_cases: list[UATTestCase], source_filename: str) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UAT Test Cases"

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    for row_idx, tc in enumerate(test_cases, start=2):
        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            value = getattr(tc, field)
            if field == "steps":
                value = "\n".join(f"{i+1}. {s}" for i, s in enumerate(value))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGNMENT

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer